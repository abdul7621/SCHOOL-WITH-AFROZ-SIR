import io
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import AppException
from app.modules.students.models import Student, Parent, StudentEnrollment
from app.modules.academics.models import ClassLevel, Section, AcademicYear
from app.modules.lookups.models import StudentStatus, LookupValue
from app.modules.fees.models import StudentFeeDemand, FeeHead, FeeInstallmentSchedule
from app.modules.excel_engine.schemas import ExcelDryRunResponse, DryRunValidationError


class ExcelMigrationService:
    @staticmethod
    def generate_student_import_template() -> bytes:
        """
        Creates a styled, standardized .xlsx template for Bulk Student Admissions.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students_Import_Template"

        # Headers
        headers = [
            "Admission_No (Required)",
            "First_Name (Required)",
            "Last_Name",
            "Gender (MALE/FEMALE/OTHER)",
            "Date_Of_Birth (YYYY-MM-DD)",
            "Class_Name (Required e.g. Class 1)",
            "Section_Name (Required e.g. Section A)",
            "Roll_No (Optional)",
            "Father_Name (Required)",
            "Mother_Name",
            "Primary_Phone (10-digits Required)",
            "WhatsApp_Phone",
            "Address",
            "Blood_Group (e.g. A+, B+, O+)",
        ]

        # Styling
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24

        # Sample Row
        sample_row = [
            "ADM-2026-0001",
            "Ayaan",
            "Khan",
            "MALE",
            "2015-05-14",
            "Class 5",
            "Section A",
            1,
            "Farhan Khan",
            "Shabana Khan",
            "9876543210",
            "9876543210",
            "12/A Civil Lines, City",
            "O+",
        ]
        ws.append(sample_row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    async def dry_run_student_import(
        cls,
        file_bytes: bytes,
        academic_year_id: str,
        db: AsyncSession,
    ) -> ExcelDryRunResponse:
        """
        Parses Excel, performs deep schema and database constraint validation,
        and returns a dry-run validation preview without altering the DB.
        """
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except Exception as e:
            raise AppException(f"Invalid Excel format or corrupted file: {str(e)}")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return ExcelDryRunResponse(
                total_rows=0,
                valid_rows_count=0,
                invalid_rows_count=0,
                can_proceed=False,
                errors=[DryRunValidationError(row_number=1, field="file", value="", error_message="Excel file has no data rows")],
            )

        # 1. Preload DB caches for fast validation
        cls_stmt = select(ClassLevel).options(selectinload(ClassLevel.sections))
        cls_res = await db.execute(cls_stmt)
        classes = cls_res.scalars().all()
        class_map = {c.name.strip().lower(): c for c in classes}

        adm_stmt = select(Student.admission_no)
        adm_res = await db.execute(adm_stmt)
        existing_admission_nos = set(adm_res.scalars().all())

        errors: List[DryRunValidationError] = []
        preview_data: List[Dict[str, Any]] = []
        seen_admissions_in_file = set()

        for idx, row in enumerate(rows[1:], start=2):
            if not row or all(v is None for v in row):
                continue  # Skip empty lines

            adm_no = str(row[0]).strip() if row[0] is not None else ""
            first_name = str(row[1]).strip() if row[1] is not None else ""
            last_name = str(row[2]).strip() if row[2] is not None else ""
            gender = str(row[3]).strip().upper() if row[3] is not None else "MALE"
            dob_raw = row[4]
            class_name = str(row[5]).strip() if row[5] is not None else ""
            section_name = str(row[6]).strip() if row[6] is not None else ""
            roll_no = int(row[7]) if row[7] is not None and str(row[7]).isdigit() else None
            father_name = str(row[8]).strip() if row[8] is not None else ""
            mother_name = str(row[9]).strip() if row[9] is not None else ""
            phone = str(row[10]).strip() if row[10] is not None else ""
            wa_phone = str(row[11]).strip() if row[11] is not None else phone
            address = str(row[12]).strip() if row[12] is not None else ""
            blood_group = str(row[13]).strip().upper() if row[13] is not None else ""

            row_has_error = False

            # Validations
            if not adm_no:
                errors.append(DryRunValidationError(row_number=idx, field="Admission_No", value="", error_message="Admission No is mandatory"))
                row_has_error = True
            elif adm_no in existing_admission_nos:
                errors.append(DryRunValidationError(row_number=idx, field="Admission_No", value=adm_no, error_message=f"Admission No '{adm_no}' already exists in School Database"))
                row_has_error = True
            elif adm_no in seen_admissions_in_file:
                errors.append(DryRunValidationError(row_number=idx, field="Admission_No", value=adm_no, error_message=f"Duplicate Admission No '{adm_no}' found in this Excel sheet"))
                row_has_error = True
            else:
                seen_admissions_in_file.add(adm_no)

            if not first_name:
                errors.append(DryRunValidationError(row_number=idx, field="First_Name", value="", error_message="First Name is required"))
                row_has_error = True

            if not father_name:
                errors.append(DryRunValidationError(row_number=idx, field="Father_Name", value="", error_message="Father Name is required"))
                row_has_error = True

            if not phone or len(phone) < 10:
                errors.append(DryRunValidationError(row_number=idx, field="Primary_Phone", value=phone, error_message="Valid 10-digit primary phone is required"))
                row_has_error = True

            # Class & Section Check
            matched_class = class_map.get(class_name.lower())
            matched_section = None
            if not matched_class:
                errors.append(DryRunValidationError(row_number=idx, field="Class_Name", value=class_name, error_message=f"Class '{class_name}' does not exist in Academics master"))
                row_has_error = True
            else:
                matched_section = next((s for s in matched_class.sections if s.name.strip().lower() == section_name.lower()), None)
                if not matched_section:
                    errors.append(DryRunValidationError(row_number=idx, field="Section_Name", value=section_name, error_message=f"Section '{section_name}' not found under '{matched_class.name}'"))
                    row_has_error = True

            # Parse DOB
            dob = None
            if isinstance(dob_raw, (date, datetime)):
                dob = dob_raw.date() if isinstance(dob_raw, datetime) else dob_raw
            elif isinstance(dob_raw, str) and dob_raw.strip():
                try:
                    dob = datetime.strptime(dob_raw.strip(), "%Y-%m-%d").date()
                except ValueError:
                    errors.append(DryRunValidationError(row_number=idx, field="Date_Of_Birth", value=dob_raw, error_message="DOB must be in YYYY-MM-DD format"))
                    row_has_error = True

            preview_data.append({
                "row_number": idx,
                "admission_no": adm_no,
                "full_name": f"{first_name} {last_name}".strip(),
                "class_name": class_name,
                "section_name": section_name,
                "class_id": matched_class.id if matched_class else None,
                "section_id": matched_section.id if matched_section else None,
                "father_name": father_name,
                "primary_phone": phone,
                "dob": str(dob) if dob else None,
                "gender": gender,
                "has_error": row_has_error,
            })

        total = len(preview_data)
        invalid = len(set(e.row_number for e in errors))
        valid = total - invalid

        return ExcelDryRunResponse(
            total_rows=total,
            valid_rows_count=valid,
            invalid_rows_count=invalid,
            can_proceed=(invalid == 0 and total > 0),
            errors=errors,
            preview_data=preview_data[:50],  # Return up to first 50 rows for preview
        )

    @classmethod
    async def execute_student_import_commit(
        cls,
        file_bytes: bytes,
        academic_year_id: str,
        db: AsyncSession,
    ) -> int:
        """
        Executes atomic database commit of all students in the Excel file.
        Rolls back entirely if any database violation occurs.
        """
        dry_run = await cls.dry_run_student_import(file_bytes, academic_year_id, db)
        if not dry_run.can_proceed:
            raise AppException(f"Cannot commit Excel with {dry_run.invalid_rows_count} validation errors. Fix errors first.")

        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # DB lookups
        cls_stmt = select(ClassLevel).options(selectinload(ClassLevel.sections))
        cls_res = await db.execute(cls_stmt)
        classes = cls_res.scalars().all()
        class_map = {c.name.strip().lower(): c for c in classes}

        # Active student status
        status_stmt = select(StudentStatus).where(StudentStatus.code == "ACTIVE")
        st_res = await db.execute(status_stmt)
        active_status = st_res.scalar_one_or_none()
        status_id = active_status.id if active_status else None

        imported_count = 0

        for row in rows[1:]:
            if not row or all(v is None for v in row):
                continue

            adm_no = str(row[0]).strip()
            first_name = str(row[1]).strip()
            last_name = str(row[2]).strip() if row[2] else ""
            gender = str(row[3]).strip().upper() if row[3] else "MALE"
            dob_raw = row[4]
            class_name = str(row[5]).strip()
            section_name = str(row[6]).strip()
            roll_no = int(row[7]) if row[7] and str(row[7]).isdigit() else None
            father_name = str(row[8]).strip()
            mother_name = str(row[9]).strip() if row[9] else ""
            phone = str(row[10]).strip()
            wa_phone = str(row[11]).strip() if row[11] else phone
            address = str(row[12]).strip() if row[12] else ""
            blood_group = str(row[13]).strip().upper() if row[13] else None

            # Date parsing
            dob = date(2015, 1, 1)
            if isinstance(dob_raw, (date, datetime)):
                dob = dob_raw.date() if isinstance(dob_raw, datetime) else dob_raw
            elif isinstance(dob_raw, str) and dob_raw.strip():
                dob = datetime.strptime(dob_raw.strip(), "%Y-%m-%d").date()

            matched_class = class_map[class_name.lower()]
            matched_section = next(s for s in matched_class.sections if s.name.strip().lower() == section_name.lower())

            # 1. Find or create Parent
            parent_stmt = select(Parent).where(Parent.primary_phone == phone)
            parent_res = await db.execute(parent_stmt)
            parent = parent_res.scalar_one_or_none()

            if not parent:
                parent = Parent(
                    father_name=father_name,
                    mother_name=mother_name,
                    primary_phone=phone,
                    whatsapp_phone=wa_phone,
                    address=address,
                )
                db.add(parent)
                await db.flush()

            # 2. Create Student
            student = Student(
                parent_id=parent.id,
                admission_no=adm_no,
                first_name=first_name,
                last_name=last_name,
                gender=gender,
                dob=dob,
                admission_date=date.today(),
                blood_group=blood_group,
                current_status_id=status_id,
            )
            db.add(student)
            await db.flush()

            # 3. Create Enrollment
            enrollment = StudentEnrollment(
                student_id=student.id,
                academic_year_id=academic_year_id,
                class_id=matched_class.id,
                section_id=matched_section.id,
                roll_no=roll_no,
                is_active=True,
            )
            db.add(enrollment)
            imported_count += 1

        await db.commit()
        return imported_count

    @staticmethod
    def export_students_to_excel(students_data: List[Dict[str, Any]], school_name: str) -> bytes:
        """
        Exports clean, formatted .xlsx spreadsheet of enrolled students.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enrolled_Students"

        # Title
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"{school_name} — ACTIVE STUDENTS DIRECTORY"
        title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="1E40AF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Admission No", "Student Name", "Class", "Section", "Roll No", "Father Name", "Phone"]
        ws.append([])  # Row 2 empty
        ws.append(headers)  # Row 3

        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

        for st in students_data:
            ws.append([
                st.get("admission_no"),
                st.get("full_name"),
                st.get("class_name"),
                st.get("section_name"),
                st.get("roll_no") or "-",
                st.get("father_name"),
                st.get("primary_phone"),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
